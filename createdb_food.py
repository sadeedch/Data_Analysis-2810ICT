"""
Python Script to access the workbooks and create a database
Author: Sadeed Ahmad
Course: 2810ICT Software Technologies
Data Analysis and Visualisation - Trimester 2, 2019
"""

# importing the sqlite and openpyxl
import sqlite3
import openpyxl

# This method removes the quotes.
def remove_quotes(string):
    return str(string).replace("'", " ").replace('"', ' ')

# creating the database.
connection = sqlite3.connect("database.db")
cursor = connection.cursor()


# Creating the inspections table if it does not exist already.
# Columns of this table have the same name as the excel file "inspections.xlsx".
# Data types of each column have been assigned accordingly.
create_inspections = """
CREATE TABLE IF NOT EXISTS inspections (
    activity_date NUMERIC,
    employee_id TEXT,
    facility_address TEXT,
    facility_city TEXT,
    facility_id TEXT,
    facility_name TEXT,
    facility_state TEXT,
    facility_zip TEXT,
    grade TEXT,
    owner_id TEXT,
    owner_name TEXT,
    pe_description TEXT,
    program_element_pe INT,
    program_name TEXT,
    program_status TEXT,
    record_id TEXT,
    score TEXT,
    serial_number TEXT,
    service_code INT,
    service_description TEXT   
) """


# Creating the violations table if it does not exist already.
# Columns of this table have the same name as the excel file "violations.xlsx".
# Data types of each column have been assigned accordingly.
create_violations = """
CREATE TABLE IF NOT EXISTS violations (
    points INT,
    serial_number TEXT,
    violation_code TEXT,
    violation_description TEXT,
    violation status TEXT
)

"""

# Executing the queries defined above to create the corresponding tables
cursor.execute(create_inspections)
cursor.execute(create_violations)


# Defining the check to see if the the inspections table is empty
check_inspections = "SELECT * FROM inspections"
check_inspections_data = cursor.execute(check_inspections).fetchall()

# Applying the above defined check, if the inspections table is empty then
# the following code will be executed
if check_inspections_data == []:

    # loading the workbook inspections.xlsx and worksheet inspections
    inspect_wb = openpyxl.load_workbook('inspections.xlsx')
    inspect_ws = inspect_wb['inspections']

    # buliding the query to insert data into inspections table columns
    inspections_data_insert = """INSERT INTO inspections VALUES"""
    for row in inspect_ws.iter_rows(min_row=2, max_row=191372):  # maximum number of rows are  191372 for this worksheet.
        inspections_data_insert += """("{}", "{}", "{}", "{}", "{}", "{}", "{}", "{}", "{}", "{}", "{}", "{}", {}, "{}", "{}", "{}", "{}", "{}", {}, "{}"), """.format(
                remove_quotes(row[0].value), remove_quotes(row[1].value),
                remove_quotes(row[2].value), remove_quotes(row[3].value),
                remove_quotes(row[4].value), remove_quotes(row[5].value),
                remove_quotes(row[6].value), remove_quotes(row[7].value),
                remove_quotes(row[8].value), remove_quotes(row[9].value),
                remove_quotes(row[10].value), remove_quotes(row[11].value),
                remove_quotes(row[12].value), remove_quotes(row[13].value),
                remove_quotes(row[14].value), remove_quotes(row[15].value),
                remove_quotes(row[16].value), remove_quotes(row[17].value),
                remove_quotes(row[18].value), remove_quotes(row[19].value),

        )

    inspections_data_insert = inspections_data_insert[:-2]

    cursor.execute(inspections_data_insert)  # executing the query


# Defining the check to see if the the violations table is empty
check_violations = "SELECT * FROM violations"
check_violations_data = cursor.execute(check_violations).fetchall()

# Applying the above defined check, if the violations table is empty then
# the following code will be executed
if check_violations_data == []:

    # loading the workbook violations.xlsx and worksheet violations
    violate_wb = openpyxl.load_workbook('violations.xlsx')
    violate_ws = violate_wb['violations']

    # buliding the query to insert data into violations table columns
    violations_data_insert = """INSERT INTO violations VALUES"""
    for row in violate_ws.iter_rows(min_row=2, max_row=906015): # maximum number of rows are  906015 for this worksheet.
        violations_data_insert += """("{}", "{}", "{}", "{}", "{}"), """.format(
                remove_quotes(row[0].value), remove_quotes(row[1].value),
                remove_quotes(row[2].value), remove_quotes(row[3].value),
                remove_quotes(row[4].value)
            )
    violations_data_insert = violations_data_insert[:-2]

    cursor.execute(violations_data_insert)  # executing the query


# commit and close the connection
connection.commit()
connection.close()